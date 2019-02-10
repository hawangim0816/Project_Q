import discord
import asyncio
import random
import openpyxl
import datetime
import os

client = discord.Client()

@client.event
async def on_ready():
    print("login")
    print(client.user.name)
    print(client.user.id)
    print("--------------")
    await client.change_presence(game=discord.Game(name='"쿼트야 도움" ㄱㄱ', type=1))
    

@client.event
async def on_message(message):
    if message.content.startswith("쿼트야 도움"):
        await client.send_message(message.channel, "**[쿼트봇 도움말]**\n"
                                            "모든 명령어는 **'쿼트야'**로 시작합니다. \n(※예외도 있음)\n"
                                            "\n"
                                            "**[기본적인 명령어]**\n"
                                            "\n"
                                            "1. 쿼트야 안녕\n"
                                            "2. 쿼트야 넌 누구야\n"
                                            "3. 주사위 3d6\n"
                                            "4. S! Expesite\n"
                                            "\n"
                                            "**[기타 명령어]**\n"
                                            "\n"
                                            "1. 간접경고 기능\n"
                                            "2. 쿼트야 팀 나눠\n"
                                            "3. 쿼트 호출하기 (쿼트야!)\n"
                                            "4. **동전던지기**\n"
                                            "\n"
                                            "그 외에도 쿼트봇을 알고싶거나 문제가 있으면\n"
                                            "**https://discord.gg/8UG9nkK** 여기와서 문의 주세요.")

    if message.content.startswith("쿼트야!"):
        await client.send_message(message.channel, "무슨 일이시죠?")

    if message.content.startswith("주사위"):
        roll = message.content.split(" ")
        rolld = roll[1].split("d")
        dice = 0
        for i in range(1, int(rolld[0]) +1):
            dice = dice + random.randint(1, int(rolld[1]))
        await client.send_message(message.channel, str(dice))

    if message.content.startswith("동전던지기"):
        chat = "앞면! 뒷면~!"
        chatchoice = chat.split(" ")
        chatnumber = random.randint(1, len(chatchoice))
        chatresult = chatchoice[chatnumber - 1]
        await client.send_message(message.channel, chatresult)

    if message.content.startswith("쿼트야 안녕"):
        file = openpyxl.load_workbook("쿨타임(도움).xlsx")
        sheet = file.active
        for i in range(1, 21):
            if sheet["A" + str(i)].value == message.author.id:
                if int(sheet["B" + str(i)].value) <= int(datetime.datetime.today().strftime("%Y%m%d%H%M%S")):
                    await client.send_message(message.channel, "안녕하세요. 엑스퍼트의 1호봇, 쿼트봇입니다.")
                    a = datetime.datetime.today() + datetime.timedelta(seconds=2)
                    sheet["B" + str(i)].value = a.strftime("%Y%m%d%H%M%S")
                    file.save("쿨타임(도움).xlsx")
                else:
                    await client.send_message(message.channel, "쿨타임은 2초입니다. 기다려주세요.")
                break
            if sheet["A" + str(i)].value =="-":
                sheet["A" + str(i)].value = message.author.id
                a = datetime.datetime.today() + datetime.timedelta(seconds=5)
                sheet["B" + str(i)].value = a.strftime("%Y%m%d%H%M%S")
                file.save("쿨타임(도움).xlsx")
                await client.send_message(message.channel, "안녕하세요. 엑스퍼트의 1호봇, 쿼트봇입니다.")
                break

    if message.content.startswith("S! Expesite"):
            await client.send_message(message.channel, "**https://suniltion.wixsite.com/teamexpert**")

    if message.content.startswith("S!link"):
            await client.send_message(message.channel, "**https://invite.gg/teamexpert**")

    if message.content.startswith("S! 삿하"):
            await client.send_message(message.channel, "삿하!")

    if message.content.startswith("쿼트야 시간"):
        now = datetime.datetime.now()
        now2 = str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second)

        print(now)
        print(now2)

        await client.send_message(message.channel, now)
        await client.send_message(message.channel, now2)

    if message.content.startswith("S! get out"):
        chat = "nope shut-up. sry-u-aren't-EPT. I-can't-get-out."
        chatchoice = chat.split(" ")
        chatnumber = random.randint(1, len(chatchoice))
        chatresult = chatchoice[chatnumber - 1]
        await client.send_message(message.channel, chatresult)

    if message.content.startswith("S! 나가"):
        chat = "못나가 뭐때문에나가야하죠? 당신은EPT가아닙니다... 10,000,000,000,000,000(1경)드리면나갈게요."
        chatchoice = chat.split(" ")
        chatnumber = random.randint(1, len(chatchoice))
        chatresult = chatchoice[chatnumber - 1]
        await client.send_message(message.channel, chatresult)

    if message.content.startswith("쿼트야 뭐해?"):
        chat = "귀찮은거? 몰라요.. 흠... hmm idk 할게없네요... 밥먹는중 장보고옴 심심해요"
        chatchoice = chat.split(" ")
        chatnumber = random.randint(1, len(chatchoice))
        chatresult = chatchoice[chatnumber - 1]
        await client.send_message(message.channel, chatresult)

    if message.content.startswith("쿼트야 ㅎㅇ"):
        chat = "ㅎㅇ 그래ㅎㅇ 반갑습니다. 그만ㅡㅡ 내가-인사머신이냐-그만해..... 귀찮지만-받아줌ㅇㅇ ㅡ_ㅡ"
        chatchoice = chat.split(" ")
        chatnumber = random.randint(1, len(chatchoice))
        chatresult = chatchoice[chatnumber - 1]
        await client.send_message(message.channel, chatresult)

    if message.content.startswith("쿼트야 넌 누구야"):
        await client.send_message(message.channel, "**쿼트야 안녕**을 입력해보세요.")

    if message.content.startswith("쿼트야 죽어"):
        await client.send_message(message.channel, "싫은데요!")

    if message.content.startswith("골라"):
        choice = message.content.split(" ")
        choicenumber = random.randint(1, len(choice) - 1)
        choiceresult = choice[choicenumber]
        await client.send_message(message.channel, choiceresult)

    if message.content.startswith("뭐 할까?"):
        chat = "뒤져 어쩌라고ㅡㅡ 내가상담원이냐?고만물어ㅡㅡ **죽어** 같이코드작업할래? 밥먹어임마 니알아서해 **꺼져** "
        chatchoice = chat.split(" ")
        chatnumber = random.randint(1, len(chatchoice))
        chatresult = chatchoice[chatnumber-1]
        await client.send_message(message.channel, chatresult)

    if message.content.startswith("뭐 먹을까?"):
        food = "아무거나 몰라 알아서!"
        foodchoice = food.split(" ")
        foodnumber = random.randint(1, len(foodchoice))
        foodresult = foodchoice[foodnumber-1]
        await client.send_message(message.channel, foodresult)

    if message.content.startswith("메모장 쓰기"):
        file = open("쿼트봇 메모장.txt", "w")
        file.write("쿼트봇에 대한 정보는 다음 기회에 공개됩니다.")
        file.close()

    if message.content.startswith("메모장 읽기"):
        file = open("쿼트봇 메모장.txt")
        await client.send_message(message.channel, file.read())
        file.close()

    if message.content.startswith("배워"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                break
        file.save("기억.xlsx")
            
    if message.content.startswith("/기억") and not message.content.startswith("기억삭제"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == memory[1]:
                await client.send_message(message.channel, sheet ["B" + str(i)].value)
                break

    if message.content.startswith("기억삭제"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = "-"
                await client.send_message(message.channel, "기억이 삭제되었으니 다시 입력하지 않으면 당신은 바보입니다.ㅋ")
                file.save("기억.xlsx")
                break

    if message.content.startswith("쿼트야 팀 나눠"):
        team = message.content[9:]
        peopleteam = team.split("/")
        people = peopleteam[0]
        team = peopleteam[1]
        person = people.split(" ")
        teamname = team.split(" ")
        random.shuffle(teamname)
        for i in range(0, len(person)):
            await client.send_message(message.channel, person[i] + "------>" + teamname[i])

    if "ㅈ까" in message.content or "애미" in message.content or "애비" in message.content or "한남" in message.content or "메갈" in message.content or "좆" in message.content or "병신" in message.content or "섹스" in message.content or "sex" in message.content:
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        member = discord.utils.get(client.get_all_members(), id="")
        for i in range(1, 251):
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                if int(sheet["B" + str(i)].value) == 10:
                    await client.ban(member, 1)
                break
            if str(sheet["A" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 1
                break
        file.save("경고.xlsx")
        await client.send_message(message.channel, "경고 1회가 부여되었습니다. 단어 사용에 주의해주세요.")

    if message.content.startswith("Q!경고"):
        memid = message.content.split(" ")
        file = openpyxl.load_workbook("경고.xlsx")
        member = discord.utils.get(client.get_all_members(), id=memid[1])
        sheet = file.active
        for i in range(1,251):
            if str(sheet["A"+ str(i)].value) == str(message.author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value + 1)
                if int(sheet["B" + str(i)].value) == 8:
                    await client.ban(member, 1)
                break
            if str(sheet["A" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(memid[1])
                sheet["B" + str(i)].value = 1
                break
        file.save("경고.xlsx")
        await client.send_message(message.channel, "경고를 받으셨습니다. 단어 사용에 주의해주세요.")

    if message.content.startswith("/역할설정"):
        role = ""
        rolename = message.content.split(" ")
        member = discord.utils.get(client.get_all_members(), id=rolename[1])
        for i in message.server.roles:
            if i.name == rolename[2]:
                role = i
                break
        await client.add_roles(member, role)


access_token = os.environ["BOT.TOKEN"]
client.run(access_token)
